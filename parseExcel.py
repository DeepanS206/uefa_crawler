from openpyxl.styles import PatternFill
from Levenshtein import distance
from openpyxl import load_workbook
from gameInfo import getMatchStats
import pickle
import sys
import os

#indices containing the games that need to be parsed, 
#passed in as command-line arguments
startInx = sys.argv[1]
endInx = sys.argv[2]
indices = startInx + ':' + endInx

#certain teams in the dataset with different naming conventions
#as teams in dataset are named in German
#resolved for the most part by using the edit-distance algorithm for team name comparison
specialTeams = {
  "Brügge":"Club_Brügge",
  "Dynamo Kiew":"Dynamo_Kyiv",
  "AC Florenz":"Fiorentina",
  "SSC Neapel":"Napoli",
  "Dynamo Moskau":"Dinamo_Moskva",
  "Aalborg BK":"AaB",
  "Young Boys Bern":"Young_Boys",
  "Dinamo Zagreb":"Dinamo_Zagreb",
  "Maccabi Haifa":" M._Haifa",
  "Pacos de Ferreira":"Paços_Ferreira",
  "Tschornomorez Odessa":"Chornomorets_Odesa",
  "Rapid Wien":"Rapid_Wien",
  "Viktoria Pilsen":"Plzeň",
  "Schachtar Donezk":"Shakhtar_Donetsk",
  "Juventus Turin":"Juventus",
  "Benfica Lissabon":"Benfica",
  "Dinamo Minsk":"Dinamo_Minsk",
  "Slovan Bratislava":"Slovan_Bratislava",
  "Rio Ave":"Rio_Ave",
  "AS St. Etienne":"St-Étienne",
  "Inter Mailand":"Internazionale",
  "Astra Giurgiu":"Astra__",
  "HNK Rijeka":"Rijeka",
  "Rubin Kasan":"Rubin__",
  "Bayer 04 Leverkusen":"Leverkusen",
  "RSC Anderlecht":"Anderlecht",
  "Manchester City":"Man._City",
  "Sporting Lissabon":"Sporting_CP",
  "Lazio":"Lazio__",
  "Kopenhagen":"København",
  "Standard Lüttich":"Standard_Liège",
  "Stade Rennes":"Rennes",
  "AEK Athen":"AEK__",
  "Atletico Madrid":"Atletico_",
  "Aalborg BK":"AaB__"
}

countryIds = {}

#loading countryIds into conutryIds dictionary
with open('countryIds.pickle', 'rb') as handle:
  countryIds = pickle.load(handle)

countryIds[''] = ''

#function to check if referee information in crawled data matches
#referee information already existing in dataset. 
#This is used as a measure of making sure the right match info is scraped
def refCheck(excel_ref, ref):
  excelArr = excel_ref.split(' ')
  refArr = ref.split(' ')
  if len(refArr) > len(excelArr):
    temp = refArr
    refArr = excelArr
    excelArr = temp
  for elem in refArr:
    tempBool = False
    for token in excelArr:
      dist = distance(elem, token)
      if dist < 2:
        tempBool = True
    if not tempBool:
      return False
  return True

#add crawled match info into corresponding cell in excel dataset
def addInfoToCell(row, matchInfo):
  redFill = PatternFill(start_color='ff1a1a', end_color='ff1a1a', fill_type='solid')
  refId = countryIds[matchInfo['ref_nat']]
  refnat = matchInfo['ref_nat']
  ycards_home = matchInfo['yellow_home']
  ycards_away = matchInfo['yellow_away']
  rcards_home = matchInfo['red_home']
  rcards_away = matchInfo['red_away']
  asstRef1nat = matchInfo['asstref1_nat']
  asstRef1id = countryIds[asstRef1nat]
  asstRef2nat = matchInfo['asstref2_nat']
  asstRef2id = countryIds[asstRef2nat]
  ref = matchInfo['ref']
  penalties = matchInfo['penalties']
  extraTimeScore = ''
  penaltyScore = ''
  if matchInfo['extra_time']:
    finalInfo = row[9].value.strip()
    arr = finalInfo.split(' ')
    if 'i.E.' in finalInfo:
      length = len(arr[0])
      penaltyScore = arr[0][4:length]
      length2 = len(arr[1])
      extraTimeScore = arr[2]
    else:
      length2 = len(finalInfo)
      extraTimeScore = finalInfo[5:length2]
  print('adding info to excel sheet')
  row[10].value = extraTimeScore
  row[11].value = penaltyScore
  row[12].value = ycards_home 
  row[13].value = ycards_away
  row[15].value = rcards_home
  row[16].value = rcards_away
  if penalties == '':
    row[20].fill = redFill
  else:
    penarr = penalties.split(':')
    row[18].value = penarr[0]
    row[19].value = penarr[1]

  if ref == '':
    row[16].fill = redFill
    row[17].fill = redFill
    row[19].fill = redFill
    row[20].fill = redFill
    row[22].fill = redFill
    row[23].fill = redFill
  else:
    row[22].value = refnat
    row[23].value = refId
    row[25].value = asstRef1nat
    row[26].value = asstRef1id
    row[28].value = asstRef2nat
    row[29].value = asstRef2id

#check to make sure correct match info is scraped
#if so, then function proceeds adding match info to cell
def checkStats(matchInfo, row, cellInx):
  query_team1 = row[4].value.strip()
  query_team2 = row[6].value.strip()
  match_ref = row[21].value.strip()
  match_asstref1 = row[24].value.strip()
  match_asstref2 = row[27].value.strip()
  ref = match_ref
  asstRef1 = match_asstref1
  asstRef2 = match_asstref2
  if 'ref' in matchInfo: #referee info exists in webpage
    ref = matchInfo['ref']
    asstRef1 = matchInfo['asstref1']
    asstRef2 = matchInfo['asstref2']
  else: #referee info does not exist in webpage
    matchInfo['ref'] = ''
    matchInfo['ref_nat'] = ''
    matchInfo['asstref1_nat'] = ''
    matchInfo['asstref2_nat'] = ''
  team1 = matchInfo['team1']
  team2 = matchInfo['team2']
  if (not refCheck(match_ref, ref)) & (not refCheck(match_asstref1, asstRef1)) & (not refCheck(match_asstref2, asstRef2)):
    print('error in matching refs, wrong game may have been found')
    print('date: ' + date + ', team1: ' + query_team1 + ', team2: ' + query_team2 + ', index: ' + str(cellInx))
    print('matched team1: ' + team1 + ', matched team2: ' + team2)
    print('keyword1: ' + word + ', keyword2: ' + word2)
    print('ref: ' + ref + ', asstref1: ' + asstRef1 + ', asstref2: ' + asstRef2)
    print('act_ref: ' + match_ref + ', act_asstref1: ' + match_asstref1 + ', act_asstref2: ' + match_asstref2)
    print('1: proceed adding info')
    print('2: skip cell')
    print('3: redo parse')
    print('4: stop')
    val = input('Selection? ')
    if val == '3':
     filter = matchInfo['filter']
     filter.append(team1)
     filter.append(team2)
     matchInfo2 = getMatchStats(date, season, group, word, word2, filter, 'europaleague')
     if matchInfo2:
      checkStats(matchInfo2, row, cellInx)
     else:
      return
    elif val == '4':
     sys.exit(1)
    elif val == '1':
     addInfoToCell(row, matchInfo)
     wb.save('Referee Assistants.xlsx')
  else:
   addInfoToCell(row, matchInfo)
   wb.save('Referee Assistants.xlsx')

cellInx = int(startInx[1:len(startInx)])
wb = load_workbook(filename='/Users/deepansaravanan/Desktop/Python/web_scraping/Referee Assistants.xlsx')
ws = wb.get_sheet_by_name('Europa League bzw. UEFA-Cup')

for row in ws.iter_rows(indices):
  date = row[0].value.strftime('%m/%d/%Y')
  season = '20' + row[1].value.strip().split('/')[1]
  group = row[2].value.strip()
  if 'Gr.,' in group:
    group = 'group'
  if 'Rd.' in group:
    group = 'Rd'
  team1 = row[4].value.strip()
  team2 = row[6].value.strip()
  if team1 in specialTeams:
    team1 = specialTeams[team1]
  if team2 in specialTeams:
    team2 = specialTeams[team2]
  team1KeyWords = team1.split(' ')
  team2KeyWords = team2.split(' ')
  print(cellInx)
  print(date)
  print(season)
  print(group)
  print(team1)
  print(team2)
  infoFound = False
  for word in team1KeyWords:
    for word2 in team2KeyWords:
      if (not infoFound) &( len(word) > 2) & (len(word2) > 2):
        #instigating crawler with the required parameters:
          #date: date of match, season: season of match, group: group of match
          #word: token of team1's name, word2: token of team2's name
          #[] filitering array, '***league': the league in which the match happened
         matchInfo = getMatchStats(date, season, group, word, word2, [], 'europaleague')
         if matchInfo:
           infoFound = True
           #if non-trivial match-info is returned, then check for correctness and 
           #add to excel file
           checkStats(matchInfo, row, cellInx)
  cellInx = cellInx + 1