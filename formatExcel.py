from openpyxl import load_workbook
from openpyxl import *
import sys

startInx = sys.argv[1]
endInx = sys.argv[2]
indices = startInx + ':' + endInx

wb = load_workbook(filename='/Users/deepansaravanan/Desktop/Python/web_scraping/Referee Assistants.xlsx')
ws = wb.get_sheet_by_name('Europa League bzw. UEFA-Cup')
for row in ws.iter_rows(indices):
  ycard_info = row[14].value
  rcard_info = row[17].value
  penalty_info = row[20].value
  game_info = row[9].value.strip()
  extraTimeScore = ''
  penaltyScore = ''
  if 'i.E.' in game_info:
    arr = game_info.split(' ')
    length = len(arr[0])
    penaltyScore = arr[0][4:length]
    length2 = len(arr[1])
    extraTimeScore = arr[2]
  elif 'n.V.' in game_info:
    length2 = len(game_info)
    extraTimeScore = game_info[5:length2]

  if (not (ycard_info is None)) & (not (rcard_info is None)) & (not (penalty_info is None)):
    arrYcards = ycard_info.strip().split(':')
    arrRcards = rcard_info.strip().split(':')
    penArr = penalty_info.strip().split(':')
    row[10].value = extraTimeScore
    row[11].value = penaltyScore
    row[12].value = arrYcards[0]
    row[13].value = arrYcards[1]
    row[15].value = arrRcards[0]
    row[16].value = arrRcards[1]
    row[18].value = penArr[0]
    row[19].value = penArr[1]
    wb.save('Referee Assistants.xlsx')
